from flask import Flask, jsonify, request, send_file, make_response
import psycopg2
from copy import copy
import datetime
from dateutil.relativedelta import relativedelta
import os
import shutil
import openpyxl
import pandas as pd
from pykakasi import kakasi
import re
import unicodedata
import traceback
import jwt
from jwt import PyJWKClient
from datetime import datetime

app = Flask(__name__)
app.json.sort_keys = False

TENANT_ID = "7e80b39f-2bf1-4395-a356-64b74b4015bb"
CLIENT_ID = "b96bf6d0-b9b0-4888-a294-83018fd7786d"
JWKS_URL = f"https://login.microsoftonline.com/{TENANT_ID}/discovery/v2.0/keys"
jwks_client = PyJWKClient(JWKS_URL)

#CORS
@app.after_request
def after_request(response):
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

# SERP DB - PostgreSQL接続
def get_connection():
    return psycopg2.connect('postgresql://serp:serp@postgres:5432/serp')

# カーソルをカラム名付き辞書形式に変換する
# カーソルはオープン状態で渡すこと
# 関数内で全行フェッチして辞書を作ります
def convertCursorToDict(cur):
    columns = [col[0] for col in cur.description]
    data = cur.fetchall()
    data_with_column_name = []
    for result in data:
        data_with_column_name.append(dict(zip(columns, result)))

    return data_with_column_name

# API No.1ファイルアップロード
@app.route("/serp/api/fileupload", methods=["POST"])
def fileupload():
    # エラーチェック
    if 'uploadFile' not in request.files:
        return jsonify({"status":-1, 'result':'uploadFile is required.'})

    file = request.files['uploadFile']
    fileName = file.filename
    if fileName == '' :
        return jsonify({"status":-1, 'result':'filename must not empty.'})

    # modified_user（Azure AD ID）を取得
    modified_user_azure_id = request.form.get("modified_user")
    if not modified_user_azure_id:
        return jsonify({"status": -1, "result": "modified_user is required."})

    conn = get_connection()
    try:
        # 共通関数で user_id を取得
        modified_user_id = get_user_id(conn, modified_user_azure_id)
        if not modified_user_id:
            return jsonify({"status": -1, "result": "User not found for azure_ad_id."})

        return jsonify({"status":_fileupload(file, modified_user_id)})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)
    finally:
        conn.close()

# API No.2 ファイル一覧
@app.route("/serp/api/filelist/<yyyymm>", methods=["GET"])
def filelist(yyyymm : str) :
    try:
        result = _filelist(yyyymm)
        if result == []:
            return jsonify({"status":1, "result":result})
        else:
            return jsonify({"status":0, "result":result})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.3 マージ結果のファイル一覧
@app.route("/serp/api/filemergelist/<yyyymm>", methods=["GET"])
def filemergelist(yyyymm: str):
    try:
        result = _filemergelist(yyyymm)
        if result == []:
            return jsonify({"status":1, "result":result})
        else:
            return jsonify({"status":0, "result":result})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.4 データ取得
@app.route("/serp/api/filedetail/<manage_id>", methods=["GET"])
def filedetail(manage_id: str):
    try:
        return jsonify({"status": 0, "result": _filedetail(manage_id)})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.5 マージ結果のデータ取得
@app.route("/serp/api/filemergedetail/<yyyymm>,<version>", methods=["GET"])
def filemergedetail(yyyymm: str, version: str):
    try:
        return jsonify({"status": 0, "result": _filemergedetail(yyyymm, version)})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.6 ファイル削除
@app.route("/serp/api/filedelete/<manage_id>", methods=["DELETE"])
def filedelete(manage_id: str):
    try:
        return jsonify({"status": 0, "result": _filedelete(manage_id)})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.8 マージ要求
@app.route("/serp/api/filemerge", methods=["PUT"])
def filemerge():
    try:
        if request.headers['Content-Type'] != 'application/json':
            return jsonify({"status": -1, "result":"Content-Type:application/jsonを指定してください"})
        # リクエストからJSON形式でデータを取得
        # 書式エラー時は例外が発生する

        # modified_user（Azure AD ID）を取得
        modified_user_azure_id = request.json['modified_user']
        if not modified_user_azure_id:
            return jsonify({"status": -1, "result": "modified_user is required."})

        # パラメータチェック(管理ID)
        manage_ids = request.json['manage_ids']
        if not isinstance(manage_ids, list):
            return jsonify({"status": -1, "result": "管理IDは配列で指定してください"})

        if manage_ids == []:
            return jsonify({"status": -1, "result": "管理IDを指定してください"})

        if len(manage_ids) != len(list(set(manage_ids))):
            return jsonify({"status": -1, "result": "管理IDの指定が重複しています"})

        fiscal_date = ''
        file_div_result = []
        with get_connection() as conn:
            # modified_user_id を共通関数で取得
            modified_user_id = get_user_id(conn, modified_user_azure_id)
            # 管理IDの取得
            manage_ids = [str(manage_id) for manage_id in manage_ids]
            for manage_id in manage_ids:
                work_date = _getFiscalDateByManageID(conn, manage_id)
                # 管理ID存在チェック
                if work_date == '':
                    return jsonify({"status": -1, "result": "管理IDが不正です"})
                if fiscal_date == '':
                    fiscal_date = work_date
                # 勘定年月が混在していたらNG
                if fiscal_date != work_date:
                    return jsonify({"status": -1, "result": "指定された管理IDは勘定年月が混在しています"})
                file_div_result.append(_getfilediv(conn, manage_id))

        # 過去月の案件情報マスタを削除フラグ=1に更新
        _updateTopicInfo(fiscal_date, modified_user_id)

        # マージ処理
        _filemerge(manage_ids, fiscal_date, modified_user_id)

        return jsonify({"status": 0, "result":fiscal_date})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.9 ファイルダウンロード
@app.route("/serp/api/filedownload/<yyyymm>,<version>", methods=["GET"])
def filedownload(yyyymm: str, version: str):
    filepath = "./" + _filedownload(yyyymm, version)
    filename = os.path.basename(filepath)
    return send_file(filepath, as_attachment=True,
                        download_name=filename,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# API No.10 案件情報マスタデータ取得
@app.route("/serp/api/topicinfolist/<group_id_flg>,<del_disp_flg>", methods=["GET"])
def topicdetail(group_id_flg: bool, del_disp_flg: bool):
    try:
        return jsonify({"status": 0, "result": _topicdetail(group_id_flg, del_disp_flg)})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.11 案件情報マスタデータ更新
@app.route("/serp/api/topicinfoupdate", methods=["PUT"])
def topicinfoupdate():
    try:
        if not request.is_json:
            return jsonify({"status": -1, "result": "Content-Type:application/jsonを指定してください"})

        topics = request.json.get("topics")
        if not isinstance(topics, list):
            return jsonify({"status": -1, "result": "topics は配列で指定してください"})
        if not topics:
            return jsonify({"status": -1, "result": "更新対象がありません"})

        # modified_user（Azure AD ID）を取得
        modified_user_azure_id = request.json['modified_user']
        if not modified_user_azure_id:
            return jsonify({"status": -1, "result": "modified_user is required."})

        updated = 0
        with get_connection() as conn:

            # modified_user_id を共通関数で取得
            modified_user_id = get_user_id(conn, modified_user_azure_id)

            with conn.cursor() as cur:
                for t in topics:
                    order_detail = t.get("order_detail")
                    order_rowno  = t.get("order_rowno")
                    group_id     = t.get("group_id")
                    disp_seq     = t.get("disp_seq")

                    if not order_detail or not order_rowno:
                        return jsonify({"status": -1, "result": "order_detail と order_rowno は必須です"})

                    sql = "update public.m_topic_info "\
                        "   set group_id = %s "\
                        "     , disp_seq = %s "\
                        "     , modified_user = %s "\
                        "     , modified_date = current_timestamp "\
                        " where order_detail = %s "\
                        "   and order_rowno  = %s "\

                    cur.execute(sql, (group_id, disp_seq, modified_user_id, order_detail, order_rowno))
                    updated += cur.rowcount

            conn.commit()

        return jsonify({"status": 0, "result": updated})
    except Exception as e:
        traceback.print_exc()
        response = {"status": -1, "error": str(e)}
        if os.environ.get("FLASK_ENV") == "development":
            response["trace"] = traceback.format_exc()
        return jsonify(response)

# API No.12 Azure認証
@app.route("/serp/api/auth/callback", methods=["POST"])
def auth_callback():
    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return {"error": "Missing or invalid token"}, 401

    # アクセストークン取得
    access_token = auth_header.split(" ")[1]

    try:
        # トークンの署名検証
        signing_key = jwks_client.get_signing_key_from_jwt(access_token)
        decoded = jwt.decode(
            access_token,
            signing_key.key,
            algorithms=["RS256"],
            audience=f"api://{CLIENT_ID}"
        )

        azure_ad_id = decoded.get("oid")
        email = decoded.get("upn")
        displayName = decoded.get("name")

        now = datetime.now()

        # DB接続＆登録／更新処理 ---
        with get_connection() as conn:
            with conn.cursor() as cur:
                # 存在チェック
                cur.execute("SELECT id FROM m_users WHERE azure_ad_id = %s", (azure_ad_id,))
                user = cur.fetchone()

                if user:
                    # 既存ユーザ → 更新
                    cur.execute("""
                        UPDATE m_users
                        SET email = %s,
                            display_name = %s,
                            last_login_at = %s,
                            updated_at = %s
                        WHERE azure_ad_id = %s
                    """, (email, displayName, now, now, azure_ad_id))
                    message = "User updated"
                else:
                    # 新規ユーザ → 登録
                    cur.execute("""
                        INSERT INTO m_users (
                            azure_ad_id, email, display_name, role_id,
                            last_login_at, created_at, updated_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (azure_ad_id, email, displayName, 2, now, now, now))
                    message = "User inserted"
            conn.commit()

        # DB保存用のレスポンス (実際はINSERT/UPDATEする)
        return jsonify({
            "status": "ok",
            "message": message,
            "azure_ad_id": azure_ad_id,
            "email": email,
            "display_name": displayName
        })

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 401

# ファイルアップロード
def _fileupload(file : any, modified_user_id : str):
    # リクエストパラメータの受取
    fiscal_date = request.form["fiscal_date"]
    file_nm = request.form["file_nm"]
    file_div = request.form["file_div"]

    # ファイル読込
    with get_connection() as conn:
        # ファイル情報マスタ 登録
        manege_id = _insertFileInfo(conn, fiscal_date, file_div, file_nm, modified_user_id)

        # ファイル区分別登録処理
        match file_div:
            case "F":   # 完成プロジェクト

                # Excelファイル読み込み
                df = pd.read_excel(file, sheet_name=0, skipfooter=1, skiprows=1)

                # 取得内容を1行・1カラムずつ分解して登録
                for index, data in df.iterrows():
                    div_cd = data[0]                    #原価部門コード
                    orders = str(data[2]).split('-')    #受注明細＋受注行番号＋部門コード
                    order_detail = orders[0]            #受注明細
                    order_rowno = orders[1]             #受注行番号
                    project_nm = data[3]                #契約工事略名
                    customer = data[4]                  #得意先名
                    cost_material = data[5]             #材料費
                    cost_labor = data[6]                #労務費
                    cost_subcontract = data[7]          #外注費
                    cost = data[8]                      #経費
                    sales = data[10]                    #売上高

                    # 完成PJ台帳 登録
                    _insertFgFileInfo(conn, manege_id, index, div_cd, order_detail, order_rowno, project_nm, customer, cost_material, cost_labor, cost_subcontract, cost, sales, modified_user_id)
                    # 案件情報マスタ 登録
                    _insertTopicInfo(conn, order_detail, order_rowno, project_nm, customer, modified_user_id)

            case "W":   # 仕掛プロジェクト

                # Excelファイル読み込み
                df = pd.read_excel(file, sheet_name=0, skipfooter=1, skiprows=1)

                # 取得内容を1行・1カラムずつ分解して登録
                for index, data in df.iterrows():
                    div_cd = data[0]                    #原価部門コード
                    orders = str(data[2]).split('-')    #受注明細＋受注行番号＋部門コード
                    order_detail = orders[0]            #受注明細
                    order_rowno = orders[1]             #受注行番号
                    project_nm = data[3]                #契約工事略名
                    customer = data[4]                  #得意先名
                    cost_material = data[5]             #材料費
                    cost_labor = data[6]                #労務費
                    cost_subcontract = data[7]          #外注費
                    cost = data[8]                      #経費

                    # 仕掛PJ台帳 登録
                    _insertWipFileInfo(conn, manege_id, index, div_cd, order_detail, order_rowno, project_nm, customer, cost_material, cost_labor, cost_subcontract, cost, modified_user_id)
                    # 案件情報マスタ 登録
                    _insertTopicInfo(conn, order_detail, order_rowno, project_nm, customer, modified_user_id)

    return 0

# 新規管理ID採番
def _getNextManageID(conn : any):
    with conn.cursor() as cur:
        cur.execute(
            "select coalesce(cast(max(cast(manage_id as numeric)) + 1 as varchar(23)), '1') as manage_id from m_file_info")
        res = cur.fetchone()
        if res is None:
            return '1'
        return res[0]

# 新規バージョン採番
def _getNextVersion(conn : any, fiscal_date : str, file_div : str):
    with conn.cursor() as cur:
        cur.execute(
            "select coalesce(cast(max(cast(version as numeric)) + 1 as varchar(2)), '1') as version from m_file_info where fiscal_date = %s and file_div = %s", (fiscal_date, file_div))
        res = cur.fetchone()
        if res is None:
            return '0'
        return res[0]

# ファイル情報マスタ登録
def _insertFileInfo(conn : any, fiscal_date : str, file_div : str, file_name : str, modified_user_id : str):
    manage_id = _getNextManageID(conn)
    version = _getNextVersion(conn, fiscal_date, file_div)

    with conn.cursor() as cur:
        cur.execute("insert into m_file_info (manage_id, fiscal_date, version, file_div, file_nm, modified_user, modified_date) values (%s, %s, %s, %s, %s, %s, current_timestamp)", (manage_id, fiscal_date, version, file_div, file_name, modified_user_id))

    return manage_id

# 全角英数字を半角に変換
def _to_half_width(text: str) -> str:
    return ''.join(unicodedata.normalize("NFKC", ch) for ch in text)

# 辞書補正（特殊ケース）
SPECIAL_MAP = {
    "ｴｽｼｰｱｲ": "SC",
    "エスシーアイ": "SC",
    # 他にも必要な特殊社名をここに追加
}

# 得意先名から英字2桁＋000を採番
def _generate_group_id_from_customer(customer: str) -> str:
    # 1. 全角→半角統一（英数字・カタカナ・記号）
    customer = unicodedata.normalize("NFKC", customer)

    # 2. 会社種別を除去
    customer = re.sub(r'^(【社内処理】株式会社|株式会社|有限会社|合同会社|合名会社|合資会社)', '', customer)

    # 3. 辞書補正
    if customer in SPECIAL_MAP:
        base = SPECIAL_MAP[customer]
    else:
        # 4. pykakasi でローマ字変換
        kks = kakasi()
        kks.setMode("J", "a")  # 漢字 → ascii
        kks.setMode("H", "a")  # ひらがな → ascii
        kks.setMode("K", "a")  # カタカナ → ascii
        kks.setMode("r", "Hepburn")  # ローマ字方式
        converter = kks.getConverter()
        romaji = converter.do(customer)

        # 5. 英字抽出
        letters = re.findall(r'[A-Za-z]', romaji.upper())

        # 6. 先頭2文字取得（不足時は補完）
        if len(letters) >= 2:
            base = letters[0] + letters[1]
        elif len(letters) == 1:
            base = letters[0] + "X"
        else:
            base = "XX"

    # 7. "000" を末尾につける
    return base + "000"

# 案件情報マスタ登録
def _insertTopicInfo(conn : any, order_detail : str, order_rowno : str, project_nm : str, customer : str, modified_user_id : str):
    with conn.cursor() as cur:
        # 受注明細・受注行番号による存在チェック
        cur.execute("""
            select 1 from m_topic_info
            where order_detail = %s and order_rowno = %s
        """, (order_detail, order_rowno))
        if cur.fetchone():
            return  # 既に存在するので何もしない

        # customer から既存の group_id を取得
        cur.execute("""
            select group_id
            from m_topic_info
            where customer = %s
            order by group_id
            limit 1
        """, (customer,))
        row = cur.fetchone()

        if row:
            group_id = row[0]
        else:
            group_id = _generate_group_id_from_customer(customer)

        # disp_seq を算出
        cur.execute("""
            select coalesce(max(disp_seq), 0) + 1
            from m_topic_info
            where group_id = %s
        """, (group_id,))
        disp_seq = cur.fetchone()[0]

        # INSERT
        cur.execute("""
            insert into m_topic_info
                (order_detail, order_rowno, project_nm, customer, group_id, disp_seq, del_flg, modified_user, modified_date)
            values (%s, %s, %s, %s, %s, %s, '0', %s, current_timestamp)
        """, (order_detail, order_rowno, project_nm, customer, group_id, disp_seq, modified_user_id))

# 完成PJ台帳登録
def _insertFgFileInfo(conn : any, manage_id : str, row_no : str, div_cd : str, order_detail : str, order_rowno : str, project_nm : str, customer : str, cost_material : str, cost_labor : str, cost_subcontract : str, cost : str, sales : str, modified_user_id : str):
    with conn.cursor() as cur:
        cur.execute("insert into t_fg_project_info (manage_id, row_no, div_cd, order_detail, order_rowno, project_nm, customer, cost_material, cost_labor, cost_subcontract, cost, sales, modified_user, modified_date) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, current_timestamp)", (manage_id, row_no, div_cd, order_detail, order_rowno, project_nm, customer, cost_material, cost_labor, cost_subcontract, cost, sales, modified_user_id))

# 仕掛PJ台帳登録
def _insertWipFileInfo(conn : any, manage_id : str, row_no : str, div_cd : str, order_detail : str, order_rowno : str, project_nm : str,customer : str, cost_material : str, cost_labor : str, cost_subcontract : str, cost : str, modified_user_id : str):
    with conn.cursor() as cur:
        cur.execute("insert into t_wip_project_info (manage_id, row_no, div_cd, order_detail, order_rowno, project_nm, customer, cost_material, cost_labor, cost_subcontract, cost, modified_user, modified_date) values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, current_timestamp)", (manage_id, row_no, div_cd, order_detail, order_rowno, project_nm, customer, cost_material, cost_labor, cost_subcontract, cost, modified_user_id))

# ファイル情報マスタから勘定年月を指定して取得
def _filelist(yyyymm: str):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'select * from m_file_info where fiscal_date = %s', (yyyymm, ))
            return convertCursorToDict(cur)

# 仕掛情報テーブルから勘定年月を指定して取得
def _wiplist(yyyymm: str):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'select * from t_wip_info where fiscal_date = %s', (yyyymm, ))
            return convertCursorToDict(cur)

# マージ結果テーブルから勘定年月を指定して取得
def _filemergelist(yyyymm: str):
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                'select t.fiscal_date, t.version, t.fg_id, f.file_nm as fg_file_name, t.wip_id, w.file_nm as wip_file_name from t_merge_target t left join m_file_info f on t.fg_id = f.manage_id left join m_file_info w on t.wip_id = w.manage_id where t.fiscal_date = %s order by version', (yyyymm, ))
            return convertCursorToDict(cur)

# データ取得
def _filedetail(manage_id: str):
    query = {
        'F': 'select f.div_cd AS "原価部門コード", md.div_nm as "原価部門名", f.order_detail as "受注明細", f.order_rowno as "受注行番号", mti.project_nm as "契約工事略名", f.customer as "得意先名", coalesce(cost_material, 0) as "材料費", coalesce(cost_labor, 0) as "労務費", coalesce(cost_subcontract, 0) as "外注費", coalesce(cost, 0) as "経費", coalesce(sales, 0) as "売上高" from t_fg_project_info f left join m_topic_info mti on f.order_detail = mti.order_detail left join m_div md on f.div_cd = md.div_cd where manage_id = %s',
        'W': 'select w.div_cd AS "原価部門コード", md.div_nm as "原価部門名", w.order_detail as "受注明細", w.order_rowno as "受注行番号", mti.project_nm as "契約工事略名", w.customer as "得意先名", coalesce(cost_material, 0) as "材料費", coalesce(cost_labor, 0) as "労務費", coalesce(cost_subcontract, 0) as "外注費", coalesce(cost, 0) as "経費" from t_wip_project_info w left join m_topic_info mti on w.order_detail = mti.order_detail left join m_div md on w.div_cd = md.div_cd where manage_id = %s'
    }

    with get_connection() as conn:
        with conn.cursor() as cur:
            file_div = _getfilediv(conn, manage_id)
            if file_div == []:
                return file_div
            cur.execute(query.get(file_div), (manage_id,))
            return convertCursorToDict(cur)

# データ削除
def _filedelete(manage_id: str):
    query = {
        'F': 'delete from t_fg_project_info where manage_id = %s',
        'W': 'delete from t_wip_project_info where manage_id = %s'
    }

    with get_connection() as conn:
        with conn.cursor() as cur:
            file_div = _getfilediv(conn, manage_id)
            if file_div == []:
                return file_div
            cur.execute(query.get(file_div), (manage_id,))
            cur.execute('delete from m_file_info where manage_id = %s', (manage_id,))
        conn.commit()

# ファイル区分取得
def _getfilediv(conn: any, manage_id: str):
    with conn.cursor() as cur:
        cur.execute(
            'select file_div from m_file_info where manage_id = %s', (manage_id, ))
        res = cur.fetchone()
        if res is None:
            return []
        return res[0]

# マージ結果テーブルから勘定年月とバージョンを指定して取得
def _filemergedetail(yyyymm: str, version: str):
    sql = ""\
    "select "\
    "      case "\
    "        when t.product_div = '1' "\
    "            then '' "\
    "        when t.product_div = '2' "\
    "            then '対象' "\
    "        end as 繰越対象 "\
    "    , t.order_detail as 受注明細 "\
    "    , mti.project_nm as 件名"\
    "    , coalesce(cost_labor, 0)       as 労務費 "\
    "    , coalesce(cost_subcontract, 0) as 外注費 "\
    "    , coalesce(cost, 0)             as 旅費交通費 "\
    "    , coalesce(cost_other, 0)       as その他 "\
    "    , coalesce(change_value, 0)     as 翌月繰越 "\
    "from "\
    "    t_merge_result t "\
    "    left join m_topic_info mti "\
    "        on t.order_detail = mti.order_detail "\
    "       and t.order_rowno = mti.order_rowno "\
    "where "\
    "    t.fiscal_date = %s "\
    "    and t.version = %s"

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql, (yyyymm, version))
            return convertCursorToDict(cur)

# 案件情報マスタ登録
def _updateTopicInfo(fiscal_date : str, modified_user_id : str):
    sql = ""\
    "update m_topic_info "\
    "set del_flg = '1' "\
    "    , modified_user = %s "\
    "    , modified_date = current_timestamp "\
    "where "\
    "    del_flg = '0' "\
    "    and (order_detail, order_rowno) in (select "\
    "                                            t_fg_project_info.order_detail "\
    "                                            , t_fg_project_info.order_rowno "\
    "                                        from "\
    "                                            t_merge_target "\
    "                                        inner join t_fg_project_info "\
    "                                            on t_fg_project_info.manage_id = t_merge_target.fg_id "\
    "                                        where "\
    "                                            t_merge_target.fiscal_date < %s "\
    "                                            and t_fg_project_info.order_detail <> 'ZAB202400001') "\

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql, (modified_user_id, fiscal_date))

# マージ要求
def _filemerge(manage_ids : str, fiscal_date : str, modified_user_id : str):
    try:
        select_version_sql = "select version from t_merge_target where fiscal_date = %s order by version desc"
        insert_target_sql = 'insert into t_merge_target (fiscal_date, version, fg_id, wip_id, modified_user, modified_date) values (%s, %s, %s, %s, %s, current_timestamp)'

        merge_sql = \
        "insert into t_merge_result "\
        "with wip as ( "\
        "    select "\
        "        order_detail "\
        "        , order_rowno "\
        "        , sum(cost_labor) + sum(cost_subcontract) + sum(cost) + sum(cost_other) as total_cost_wip "\
        "        , sum(cost_labor) as cost_labor_wip "\
        "        , sum(cost_subcontract) as cost_subcontract_wip "\
        "        , sum(cost) as cost_wip "\
        "        , sum(cost_other) as cost_other_wip "\
        "    from "\
        "        t_wip_info "\
        "    where "\
        "        fiscal_date = %s "\
        "    group by "\
        "        order_detail, order_rowno "\
        ") "\
        "select "\
        "    fiscal_date "\
        "    , %s as version "\
        "    , t_fg_project_info.order_detail "\
        "    , t_fg_project_info.order_rowno "\
        "    , t_fg_project_info.manage_id as fg_id "\
        "    , null as wip_id "\
        "    , cost_labor - coalesce(cost_labor_wip, 0) as cost_labor "\
        "    , cost_subcontract - coalesce(cost_subcontract_wip, 0) as cost_subcontract "\
        "    , cost - coalesce(cost_wip, 0) as cost "\
        "    , cost_material - coalesce(cost_other_wip, 0) as cost_other "\
        "    , null as change_value "\
        "    , '1' as product_div "\
        "    , %s as modified_user "\
        "    , current_timestamp as modified_date "\
        "from "\
        "    t_fg_project_info "\
        "    left join wip "\
        "        on t_fg_project_info.order_detail = wip.order_detail "\
        "       and t_fg_project_info.order_rowno = wip.order_rowno "\
        "    inner join m_file_info "\
        "        on t_fg_project_info.manage_id = m_file_info.manage_id "\
        "where "\
        "    t_fg_project_info.manage_id in %s "\
        "union all "\
        "select "\
        "    fiscal_date "\
        "    , %s as version "\
        "    , t_wip_project_info.order_detail "\
        "    , t_wip_project_info.order_rowno "\
        "    , null as fg_id "\
        "    , t_wip_project_info.manage_id as wip_id "\
        "    , cost_labor - coalesce(cost_labor_wip, 0) as cost_labor "\
        "    , cost_subcontract - coalesce(cost_subcontract_wip, 0) as cost_subcontract "\
        "    , cost - coalesce(cost_wip, 0) as cost "\
        "    , cost_material - coalesce(cost_other_wip, 0) as cost_other "\
        "    , cost_labor + cost_subcontract + cost + cost_material as change_value "\
        "    , '2' as product_div "\
        "    , %s as modified_user "\
        "    , current_timestamp as modified_date "\
        "from "\
        "    t_wip_project_info "\
        "    left join wip "\
        "        on t_wip_project_info.order_detail = wip.order_detail "\
        "       and t_wip_project_info.order_rowno = wip.order_rowno "\
        "    inner join m_file_info "\
        "        on t_wip_project_info.manage_id = m_file_info.manage_id "\
        "where "\
        "    t_wip_project_info.manage_id in %s "

        with get_connection() as conn:
            # バージョン採番
            version = '0'
            with conn.cursor() as cur:
                cur.execute(select_version_sql, (fiscal_date,))
                res = cur.fetchone()
                if res is None:
                    version = '0'
                else:
                    version = str(int(res[0]) + 1)

            # マージ処理
            with conn.cursor() as cur:
                cur.execute(merge_sql, (_getPrevMonth(fiscal_date), version, modified_user_id, tuple(manage_ids), version, modified_user_id, tuple(manage_ids),))

                # 管理ID初期化
                fg = '';
                wip = '';

                # マージ対象ファイルIDを保存
                for manage_id in manage_ids:
                    file_div = str(_getfilediv(conn, manage_id))
                    match file_div:
                        case 'F':
                            fg = manage_id
                        case 'W':
                            wip = manage_id
                            # 勘定年月分の仕掛情報テーブルを洗替え
                            cur.execute("delete from t_wip_info where fiscal_date = %s", (fiscal_date,))
                            cur.execute("insert into t_wip_info (fiscal_date, order_detail, order_rowno, cost_labor, cost_subcontract, cost, cost_other, modified_user, modified_date) select %s as fiscal_date, order_detail, order_rowno, cost_labor, cost_subcontract, cost, cost_material, %s as modified_user, current_timestamp from t_wip_project_info where manage_id = %s", (fiscal_date, modified_user_id, manage_id,))
                cur.execute(insert_target_sql, (fiscal_date, version, fg, wip, modified_user_id,))
    except Exception as e:
        print("=== ERROR OCCURRED ===")
        print("エラー内容:", e)
        traceback.print_exc()
        raise

def _filedownload(yyyymm : str, version : str):
    target_file = yyyymm + '.xlsx'

    shutil.copy('template.xlsx', target_file)
    wb = openpyxl.load_workbook(target_file)

    template_sheet = wb['template']

    # =====================
    # サマリー・仕掛MM月シート
    # =====================
    beginningOfYear = _get_last_year_november(yyyymm)  # 前年11月（期首月）
    currentyyyymm = yyyymm  # 当月

    while yyyymm >=  beginningOfYear:
        ws = wb.copy_worksheet(template_sheet)
        ws.title = "サマリー・仕掛り" + yyyymm[4:6] + "月"

        # ヘッダ行
        ws.cell(1, 2, yyyymm[0:4] + "年" + yyyymm[4:6] + "月　仕掛 原価一覧")

        # サマリー
        result = _loadmerge(yyyymm, version)
        result_fg = _loadmerge_fg(yyyymm, version)
        result_wip = _loadmerge_wip(yyyymm, version)
        result_prev_wip = _loatmerge_perv_wip(yyyymm)

        ws.unmerge_cells('H8:I8')

        # データ件数に応じて行を追加
        _insert_rows_with_style(ws, 3, len(result), False)

        row = 3 # 仕掛原価一覧 開始行
        noIndirect = 0
        prev_group_id = None # 前の行のグループID
        empty_row_count = 0 # 空行の数
        for item in result:
            group_id = item['group_id']

            # グループIDが切り替わったら空行を追加（間接プロジェクト除く）
            if prev_group_id is not None and group_id != prev_group_id and item['order_detail'] != 'ZAB202400001':
                _insert_rows_with_style(ws, row, 3, False)
                ws.cell(row, 10, '=F' + str(row) + '+G' + str(row) + '+H' + str(row) + '+I' + str(row) + '')                                        # 小計
                ws.cell(row, 11, '=IF(B' + str(row) + '="○",IF(E' + str(row) + '="",0,J' + str(row) + '),D' + str(row) + '+J' + str(row) + ')')     # 振替額
                ws.cell(row, 12, '=IF(B' + str(row) + '="○",D' + str(row) + '+J' + str(row) + '-K' + str(row) + ',"--")')                           # 翌月繰越
                row += 1
                empty_row_count += 1

            # 課題No3暫定対応
            if item['order_detail'] == 'ZAB202400001':
                indirect = copy(item)
                noIndirect = 1;
                continue
            if item['product_div'] == '2':
                ws.cell(row, 2, "○")                                                                                                        # 繰越(仕掛)
            else:
                ws.cell(row, 2, "")                                                                                                         # 繰越(完成)
            ws.cell(row, 3, item['project_nm'])                                                                                             # 件名
            ws.cell(row, 4, item['total_cost_wip'])                                                                                         # 前月繰越残
            ws.cell(row, 5, item['sales'])                                                                                                  # 当月売上
            ws.cell(row, 6, item['cost_labor'])                                                                                             # 労務費
            ws.cell(row, 7, item['cost_subcontract'])                                                                                       # 外注費
            ws.cell(row, 8, item['cost'])                                                                                                   # 旅費交通費
            ws.cell(row, 9, item['cost_other'])                                                                                             # その他
            ws.cell(row, 10, '=F' + str(row) + '+G' + str(row) + '+H' + str(row) + '+I' + str(row) + '')                                    # 小計
            ws.cell(row, 11, '=IF(B' + str(row) + '="○",IF(E' + str(row) + '="",0,J' + str(row) + '),D' + str(row) + '+J' + str(row) + ')') # 振替額
            ws.cell(row, 12, '=IF(B' + str(row) + '="○",D' + str(row) + '+J' + str(row) + '-K' + str(row) + ',"--")')                       # 翌月繰越
            row += 1

            # 次のループの比較用
            prev_group_id = group_id

        ws.cell(row, 10, '=F' + str(row) + '+G' + str(row) + '+H' + str(row) + '+I' + str(row) + '')                                        # 小計
        ws.cell(row, 11, '=IF(B' + str(row) + '="○",IF(E' + str(row) + '="",0,J' + str(row) + '),D' + str(row) + '+J' + str(row) + ')')     # 振替額
        ws.cell(row, 12, '=IF(B' + str(row) + '="○",D' + str(row) + '+J' + str(row) + '-K' + str(row) + ',"--")')                           # 翌月繰越

        # 計算式を更新（間接プロジェクト）
        row += 1

        # 間接プロジェクトがある場合のみ
        if noIndirect == 1:
            ws.cell(row, 6, indirect['cost_labor'])                                                                                         # 労務費
            ws.cell(row, 7, indirect['cost_subcontract'])                                                                                   # 外注費
            ws.cell(row, 8, 0)                                                                                                              # 旅費交通費
            ws.cell(row, 9, indirect['cost'])                                                                                               # その他
            ws.cell(row, 10, '=F' + str(row) + '+G' + str(row) + '+H' + str(row) + '+I' + str(row) + '')                                    # 小計
            ws.cell(row, 11, '=IF(B' + str(row) + '="○",IF(E' + str(row) + '="",0,J' + str(row) + '),D' + str(row) + '+J' + str(row) + ')') # 振替額
            ws.cell(row, 12, '=IF(B' + str(row) + '="○",D' + str(row) + '+J' + str(row) + '-K' + str(row) + ',"--")')                       # 翌月繰越
            row += 1  # 空行

        row += 1  # 合計 開始行

        ws.cell(row, 4, '=SUM(D2:D' + str(row-2) + ')')
        ws.cell(row, 5, '=SUM(E2:E' + str(row-2) + ')')
        ws.cell(row, 6, '=SUM(F2:F' + str(row-2) + ')')
        ws.cell(row, 7, '=SUM(G2:G' + str(row-2) + ')')
        ws.cell(row, 8, '=SUM(H2:H' + str(row-2) + ')')
        ws.cell(row, 9, '=SUM(I2:I' + str(row-2) + ')')
        ws.cell(row, 10, '=SUM(J2:J' + str(row-2) + ')')
        ws.cell(row, 11, '=SUM(K2:K' + str(row-2) + ')')
        ws.cell(row, 12, '=SUM(L2:L' + str(row-2) + ')')

        # 行幅の設定
        for r in range(len(result) + empty_row_count + 3):
            ws.row_dimensions[r + 4].height = copy(ws.row_dimensions[3].height)

        # セルの結合
        ws.merge_cells('H' + str(len(result) + empty_row_count + 6) + ':I' + str(len(result) + empty_row_count + 6))

        ws.cell(len(result) + empty_row_count + 6, 8, '=H' + str(len(result) + empty_row_count + 5) + '+I' + str(len(result) + empty_row_count + 5))

        row += 7  # 完成PJ 開始行

        # ================================ 完成PJ ================================
        fg_data_num = len(result_fg) # 完成PJのデータ数
        if fg_data_num > 2 :
            # データ件数に応じて行を追加
            _insert_rows_with_style(ws, row, fg_data_num, False)

        fg_start_row = row
        for item in result_fg:
            ws.cell(row, 3, item['project_nm'])         # 件名
            ws.cell(row, 4, item['cost_material'])      # 材料費
            ws.cell(row, 5, item['cost_labor'])         # 労務費
            ws.cell(row, 6, item['cost_subcontract'])   # 外注費
            ws.cell(row, 7, item['cost'])               # 旅費交通費
            ws.cell(row, 8, item['total_cost'])         # 原価計
            ws.cell(row, 9, item['sales'])              # 売上高
            row += 1

        # 完成PJサマリ 開始行
        row = fg_start_row + fg_data_num + max(0, 2 - fg_data_num) # 固定行(2行)にデータが入らなかった場合に備えて、計算式を設定
        if fg_data_num > 0:
            ws.cell(row, 8, '=SUM(H' + str(fg_start_row) + ':H' + str(fg_start_row + fg_data_num - 1) + ')')  # 原価計 合計
            ws.cell(row, 9, '=SUM(I' + str(fg_start_row) + ':I' + str(fg_start_row + fg_data_num - 1) + ')')  # 売上高 合計

        row += 4  # 仕掛PJ 開始行

        # ================================ 仕掛PJ ================================
        wip_data_num = len(result_wip) # 仕掛PJのデータ数

        # データ件数に応じて行を追加
        if wip_data_num > 2:
            _insert_rows_with_style(ws, row, wip_data_num, False)

        wip_start_row = row
        for item in result_wip:
            ws.cell(row, 3, item['project_nm'])         # 件名
            ws.cell(row, 4, item['cost_material'])      # 材料費
            ws.cell(row, 5, item['cost_labor'])         # 労務費
            ws.cell(row, 6, item['cost_subcontract'])   # 外注費
            ws.cell(row, 7, item['cost'])               # 経費
            ws.cell(row, 8, item['total_cost'])         # 原価計
            row += 1

        #仕掛PJサマリ 開始行
        row = wip_start_row + wip_data_num + max(0, 2 - wip_data_num) #固定行(2行)にデータが入らなかった場合に備えて、計算式を設定
        if wip_data_num > 0:
            ws.cell(row, 8, '=SUM(H' + str(wip_start_row) + ':H' + str(wip_start_row + wip_data_num - 1) + ')')  # 原価計 合計

        row += 4  # 前月仕掛 開始行

        # ================================ 前月仕掛 ================================
        prev_wip_data_num = len(result_prev_wip) # 前月仕掛のデータ数

        # データ件数に応じて行を追加
        if prev_wip_data_num > 2:
            _insert_rows_with_style(ws, row, prev_wip_data_num, False)

        wip_start_row = row
        for item in result_prev_wip:
            ws.cell(row, 3, item['project_nm'])         # 件名
            ws.cell(row, 4, item['cost_other'])         # 材料費
            ws.cell(row, 5, item['cost_labor'])         # 労務費
            ws.cell(row, 6, item['cost_subcontract'])   # 外注費
            ws.cell(row, 7, item['cost'])               # 経費
            row += 1
        for index in range(prev_wip_data_num):
            ws.cell(wip_start_row + index, 8, '=SUM(D' + str(wip_start_row + index) + ':G' + str(wip_start_row + index) + ')')   # 計

        #前月仕掛サマリ 開始行
        row = wip_start_row + prev_wip_data_num + max(0, 2 - prev_wip_data_num)  # 固定行(2行)にデータが入らなかった場合に備えて、計算式を設定
        if prev_wip_data_num > 0:
            ws.cell(row, 8, '=SUM(H' + str(wip_start_row) + ':H' + str(wip_start_row + prev_wip_data_num - 1) + ')')  # 計 合計

        row += 4  #完成PJ+仕掛PJ 開始行

        # ================================ 完成PJ＋仕掛PJ 当月増加分 ================================
        fg_add_data_num = sum(1 for row in result if row['product_div'] == '1') # 当月増加分（完成PJ）のデータ数

        # データ件数に応じて行を追加
        if fg_add_data_num > 2:
            _insert_rows_with_style(ws, row, fg_add_data_num, True)

        total_fg_start_row = row
        for item in result:
            if item['product_div'] == '1':
                ws.cell(row, 3, item['project_nm'])         # 件名
                ws.cell(row, 4, item['cost_other'])         # 材料費
                ws.cell(row, 5, item['cost_labor'])         # 労務費
                ws.cell(row, 6, item['cost_subcontract'])   # 外注費
                ws.cell(row, 7, item['cost'])               # 経費
                row += 1
        row = total_fg_start_row + fg_add_data_num  # 完成PJ+仕掛PJ内仕掛　開始行

        wip_add_data_num = sum(1 for row in result if row['product_div'] == '2') # 当月増加分（仕掛PJ）のデータ数

        # データ件数に応じて行を追加
        if wip_add_data_num > 2:
            _insert_rows_with_style(ws, row, wip_add_data_num, True)

        total_wip_start_row = row
        for item in result:
            if item['product_div'] == '2':
                ws.cell(row, 3, item['project_nm'])         # 件名
                ws.cell(row, 4, item['cost_other'])         # 材料費
                ws.cell(row, 5, item['cost_labor'])         # 労務費
                ws.cell(row, 6, item['cost_subcontract'])   # 外注費
                ws.cell(row, 7, item['cost'])               # 経費
                row += 1
        for index in range(fg_add_data_num + wip_add_data_num):
            ws.cell(total_fg_start_row + index, 8, '=SUM(D' + str(total_fg_start_row + index) + ':G' + str(total_fg_start_row + index) + ')')   # 小計

        #完成PJ+仕掛PJサマリ 開始行
        row = total_wip_start_row + wip_add_data_num + max (0, 2 - fg_add_data_num) + max(0, 2 - wip_add_data_num) #固定(完成2行、仕掛2行)にデータが入らなかった場合に備えて、計算式を設定
        if fg_add_data_num + wip_add_data_num > 0:
            ws.cell(row, 4, '=SUM(D' + str(total_fg_start_row) + ':D' + str(total_fg_start_row + fg_add_data_num + wip_add_data_num - 1) + ')')  # 材料費 合計
            ws.cell(row, 5, '=SUM(E' + str(total_fg_start_row) + ':E' + str(total_fg_start_row + fg_add_data_num + wip_add_data_num - 1) + ')')  # 労務費 合計
            ws.cell(row, 6, '=SUM(F' + str(total_fg_start_row) + ':F' + str(total_fg_start_row + fg_add_data_num + wip_add_data_num - 1) + ')')  # 外注費 合計
            ws.cell(row, 7, '=SUM(G' + str(total_fg_start_row) + ':G' + str(total_fg_start_row + fg_add_data_num + wip_add_data_num - 1) + ')')  # 経費   合計
            ws.cell(row, 8, '=SUM(H' + str(total_fg_start_row) + ':H' + str(total_fg_start_row + fg_add_data_num + wip_add_data_num - 1) + ')')  # 小計   合計

        # 前月取得
        yyyymm = _getPrevMonth(yyyymm)

    # templeteシートを削除
    del wb['template']

    # サマリー
    result_fg = _loadmerge_fg(currentyyyymm, version)
    result_wip = _loadmerge_wip(currentyyyymm, version)

    # =====================
    # 完成PJ台帳シート
    # =====================
    ws = wb['完成PJ台帳']

    fg_data_num = len(result_fg) # 完成PJのデータ数
    row = 3 # 完成PJ 開始行
    fg_start_row = row

    # データ件数に応じて行を追加
    if fg_data_num > 2:
        _insert_rows_with_style(ws, row, fg_data_num, False)

    for item in result_fg:
        ws.cell(row, 1, item['div_cd'])                         # 原価部門コード
        ws.cell(row, 2, item['div_nm'])                         # 原価部門名
        ws.cell(row, 3, item['order_detail'])                   # 受注明細+受注行番号+部門コード
        ws.cell(row, 4, item['project_nm'])                     # 件名
        ws.cell(row, 5, item['customer'])                       # 得意先名
        ws.cell(row, 6, item['cost_material'])                  # 材料費
        ws.cell(row, 7, item['cost_labor'])                     # 労務費
        ws.cell(row, 8, item['cost_subcontract'])               # 外注費
        ws.cell(row, 9, item['cost'])                           # 経費
        ws.cell(row, 10, item['total_cost'])                    # 原価計
        ws.cell(row, 11, item['sales'])                         # 売上高
        ws.cell(row, 12, item['sales'] - item['total_cost'])    # 粗利
        ws.cell(row, 13, '=IF(K' + str(row) + '=0, 0%,L' + str(row) + '/K' + str(row) + ')')  # 粗利率
        row += 1
    row = fg_start_row + fg_data_num + max(0, 2 - fg_data_num) # 完成PJ総計 開始行
    if fg_data_num > 0:
        ws.cell(row, 6, '=SUM(F' + str(fg_start_row) + ':F' + str(fg_start_row + fg_data_num - 1) + ')')     # 材料費計
        ws.cell(row, 7, '=SUM(G' + str(fg_start_row) + ':G' + str(fg_start_row + fg_data_num - 1) + ')')     # 労務費計
        ws.cell(row, 8, '=SUM(H' + str(fg_start_row) + ':H' + str(fg_start_row + fg_data_num - 1) + ')')     # 外注計
        ws.cell(row, 9, '=SUM(I' + str(fg_start_row) + ':I' + str(fg_start_row + fg_data_num - 1) + ')')     # 経費計
        ws.cell(row, 10, '=SUM(J' + str(fg_start_row) + ':J' + str(fg_start_row + fg_data_num - 1) + ')')    # 原価計
        ws.cell(row, 11, '=SUM(K' + str(fg_start_row) + ':K' + str(fg_start_row + fg_data_num - 1) + ')')    # 売上高計
        ws.cell(row, 12, '=+K' + str(fg_start_row + fg_data_num) + '-J' + str(fg_start_row + fg_data_num))        # 粗利計
        ws.cell(row, 13, '=IF(K' + str(fg_start_row + fg_data_num) + '=0, 0%,L' + str(fg_start_row + fg_data_num) + '/K' + str(fg_start_row + fg_data_num) + ')')  # 粗利率

    # =====================
    # 仕掛PJ台帳シート
    # =====================
    ws = wb['仕掛PJ台帳']

    wip_data_num = len(result_wip) # 仕掛PJのデータ数
    row = 3 # 仕掛PJ 開始行
    wip_start_row = row

    # データ件数に応じて行を追加
    if wip_data_num > 2:
        _insert_rows_with_style(ws, row, wip_data_num, False)

    for item in result_wip:
        ws.cell(row, 1, item['div_cd'])             # 原価部門コード
        ws.cell(row, 2, item['div_nm'])             # 原価部門名
        ws.cell(row, 3, item['order_detail'])       # 受注明細+受注行番号+部門コード
        ws.cell(row, 4, item['project_nm'])         # 件名
        ws.cell(row, 5, item['customer'])           # 得意先名
        ws.cell(row, 6, item['cost_material'])      # 材料費
        ws.cell(row, 7, item['cost_labor'])         # 労務費
        ws.cell(row, 8, item['cost_subcontract'])   # 外注費
        ws.cell(row, 9, item['cost'])               # 経費
        ws.cell(row, 10, item['total_cost'])        # 合計
        row += 1
    row = wip_start_row + wip_data_num + max(0, 2 - wip_data_num) # 仕掛PJサマリ 開始行
    if wip_data_num > 0:
        ws.cell(row, 6, '=SUM(F' + str(fg_start_row) + ':F' + str(fg_start_row + wip_data_num - 1) + ')')      # 材料費計
        ws.cell(row, 7, '=SUM(G' + str(fg_start_row) + ':G' + str(fg_start_row + wip_data_num - 1) + ')')      # 労務費計
        ws.cell(row, 8, '=SUM(H' + str(fg_start_row) + ':H' + str(fg_start_row + wip_data_num - 1) + ')')      # 外注計
        ws.cell(row, 9, '=SUM(I' + str(fg_start_row) + ':I' + str(fg_start_row + wip_data_num - 1) + ')')      # 経費計
        ws.cell(row, 10, '=SUM(J' + str(fg_start_row) + ':J' + str(fg_start_row + wip_data_num - 1) + ')')     # 原価計

    wb.save(target_file)

    return target_file

# マージ結果
def _loadmerge(yyyymm : str, version : str):
    sql = "with wip as ( "\
        "    select "\
        "          order_detail "\
        "        , order_rowno "\
        "        , sum(cost_labor) + sum(cost_subcontract) + sum(cost) + sum(cost_other) as total_cost_wip "\
        "        , sum(cost_labor) as cost_labor_wip "\
        "        , sum(cost_subcontract) as cost_subcontract_wip "\
        "        , sum(cost) as cost_wip "\
        "        , sum(cost_other) as cost_other_wip "\
        "    from "\
        "        t_wip_info "\
        "    where "\
        "         fiscal_date = %s"\
        "    group by "\
        "        order_detail"\
        "        , order_rowno"\
        ") "\
        "select"\
        "     t_fg_project_info.order_detail "\
        "    , m_topic_info.project_nm "\
        "    , wip.total_cost_wip "\
        "    , t_fg_project_info.sales "\
        "    , t_fg_project_info.cost_labor - coalesce(wip.cost_labor_wip, 0)   as cost_labor "\
        "    , t_fg_project_info.cost_subcontract - coalesce(wip.cost_subcontract_wip, 0) as cost_subcontract "\
        "    , t_fg_project_info.cost - coalesce(wip.cost_wip, 0) as cost "\
        "    , t_fg_project_info.cost_material - coalesce(wip.cost_other_wip, 0) as cost_other "\
        "    , null as change_value "\
        "    , '1' as product_div "\
        "    , m_topic_info.group_id "\
        "    , m_topic_info.disp_seq "\
        "from "\
        "    t_fg_project_info "\
        "    left join wip "\
        "        on t_fg_project_info.order_detail = wip.order_detail "\
        "        and t_fg_project_info.order_rowno = wip.order_rowno "\
        "    left join m_topic_info "\
        "        on t_fg_project_info.order_detail = m_topic_info.order_detail "\
        "        and t_fg_project_info.order_rowno = m_topic_info.order_rowno "\
        "    inner join m_file_info "\
        "        on t_fg_project_info.manage_id = m_file_info.manage_id "\
        "where "\
        "    t_fg_project_info.manage_id in (select fg_id from t_merge_target where fiscal_date = %s and version = %s) "\
        "union all "\
        "select "\
        "     t_wip_project_info.order_detail "\
        "    , m_topic_info.project_nm "\
        "    , wip.total_cost_wip "\
        "    , null as sales "\
        "    , t_wip_project_info.cost_labor - coalesce(wip.cost_labor_wip, 0)   as cost_labor "\
        "    , t_wip_project_info.cost_subcontract - coalesce(wip.cost_subcontract_wip, 0) as cost_subcontract "\
        "    , t_wip_project_info.cost - coalesce(wip.cost_wip, 0) as cost "\
        "    , t_wip_project_info.cost_material - coalesce(wip.cost_other_wip, 0) as cost_other "\
        "    , t_wip_project_info.cost_labor + t_wip_project_info.cost_subcontract + t_wip_project_info.cost + t_wip_project_info.cost_material as change_value "\
        "    , '2' as product_div "\
        "    , m_topic_info.group_id "\
        "    , m_topic_info.disp_seq "\
        "from "\
        "    t_wip_project_info "\
        "    left join wip "\
        "        on t_wip_project_info.order_detail = wip.order_detail "\
        "    left join m_topic_info "\
        "        on t_wip_project_info.order_detail = m_topic_info.order_detail "\
        "        and t_wip_project_info.order_rowno = m_topic_info.order_rowno "\
        "    inner join m_file_info "\
        "        on t_wip_project_info.manage_id = m_file_info.manage_id "\
        "where "\
        "    t_wip_project_info.manage_id in (select wip_id from t_merge_target where fiscal_date = %s and version = %s) "\
        "order by "\
        "    group_id "\
        "    , disp_seq "\
        "    , order_detail "
    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql, (_getPrevMonth(yyyymm), yyyymm, version, yyyymm, version, ))
            return convertCursorToDict(cur)

# マージ結果（完成）
def _loadmerge_fg(yyyymm : str, version : str):
    sql = "select "\
        "      m_div.div_cd "\
        "    , m_div.div_nm "\
        "    , m_topic_info.order_detail || '-' || m_topic_info.order_rowno || '-' || m_div.div_cd as order_detail "\
        "    , m_topic_info.project_nm "\
        "    , t_fg_project_info.customer "\
        "    , t_fg_project_info.cost_material "\
        "    , t_fg_project_info.cost_labor "\
        "    , t_fg_project_info.cost_subcontract "\
        "    , t_fg_project_info.cost "\
        "    ,  t_fg_project_info.cost_material + t_fg_project_info.cost_labor + t_fg_project_info.cost_subcontract + t_fg_project_info.cost as total_cost "\
        "    , t_fg_project_info.sales "\
        "from "\
        "    t_fg_project_info "\
        "    left join m_topic_info "\
        "        on t_fg_project_info.order_detail = m_topic_info.order_detail "\
        "        and t_fg_project_info.order_rowno = m_topic_info.order_rowno "\
        "    inner join m_div "\
        "        on t_fg_project_info.div_cd = m_div.div_cd "\
        "where "\
        "    t_fg_project_info.manage_id in (select fg_id from t_merge_target where fiscal_date = %s and version = %s) "\
        "order by "\
        "    t_fg_project_info.row_no, t_fg_project_info.order_detail "

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql, (yyyymm, version, ))
            return convertCursorToDict(cur)

# マージ結果（仕掛）
def _loadmerge_wip(yyyymm : str, version : str):
    sql = "select "\
        "      m_div.div_cd "\
        "    , m_div.div_nm "\
        "    , m_topic_info.order_detail || '-' || m_topic_info.order_rowno || '-' || m_div.div_cd as order_detail "\
        "    , m_topic_info.project_nm "\
        "    , t_wip_project_info.customer "\
        "    , t_wip_project_info.cost_material "\
        "    , t_wip_project_info.cost_labor "\
        "    , t_wip_project_info.cost_subcontract "\
        "    , t_wip_project_info.cost "\
        "    , t_wip_project_info.cost_material + t_wip_project_info.cost_labor + t_wip_project_info.cost_subcontract + t_wip_project_info.cost as total_cost "\
        "from "\
        "    t_wip_project_info "\
        "    left join m_topic_info "\
        "        on t_wip_project_info.order_detail = m_topic_info.order_detail "\
        "        and t_wip_project_info.order_rowno = m_topic_info.order_rowno "\
        "    inner join m_div "\
        "        on t_wip_project_info.div_cd = m_div.div_cd "\
        "where "\
        "    t_wip_project_info.manage_id in (select wip_id from t_merge_target where fiscal_date = %s and version = %s) "\
        "order by "\
        "    t_wip_project_info.row_no, t_wip_project_info.order_detail "

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql, (yyyymm, version, ))
            return convertCursorToDict(cur)

def _loatmerge_perv_wip(yyyymm:str):
    sql = "select "\
        "      project_nm "\
        "    , cost_labor "\
        "    , cost_subcontract "\
        "    , cost "\
        "    , cost_other "\
        "from "\
        "    t_wip_info "\
        "    left join m_topic_info "\
        "        on t_wip_info.order_detail = m_topic_info.order_detail "\
        "where "\
        "    fiscal_date = %s "\
        "order by "\
        "    t_wip_info.order_detail "

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(
                sql, (_getPrevMonth(yyyymm), ))
            return convertCursorToDict(cur)

# 案件情報マスタからgroup_idの有無、削除フラグの状態を指定して取得
def _topicdetail(group_id_flg: bool, del_disp_flg: bool):
    sql = "select "\
        "      order_detail "\
        "    , order_rowno "\
        "    , project_nm "\
        "    , group_id "\
        "    , disp_seq "\
        "    , del_flg "\
        "from "\
        "    m_topic_info "\
        "where "\
        "    (%s = FALSE OR group_id IS NULL) "\
        "and (%s = TRUE OR del_flg = '0') "\
        "order by "\
        "    group_id, disp_seq "

    with get_connection() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (group_id_flg, del_disp_flg, ))
            return convertCursorToDict(cur)

# 勘定年月取得
def _getFiscalDate():
    # 前月末日を取得しその年月を文字列として返す
    today = datetime.today()
    thismonth = datetime(today.year, today.month, 1)
    lastmonth = thismonth + datetime.timedelta(days=-1)
    return lastmonth.strftime("%Y%m")

# 勘定年月取得
def _getFiscalDateByManageID(conn : any, manage_id : str):
    with conn.cursor() as cur:
        cur.execute(
            'select fiscal_date from m_file_info where manage_id = %s', (manage_id, ))
        res = cur.fetchone()
        if res is None:
            return ''
        return res[0]

# 前月取得
def _getPrevMonth(yyyymm : str):
    return (datetime.strptime(yyyymm, '%Y%m') + relativedelta(months=-1)).strftime("%Y%m")

# 期首月前月取得
def _get_last_year_november(yyyymm: str) -> str:
    # 入力（例: "202408"）を datetime に変換
    dt = datetime.strptime(yyyymm, "%Y%m")

    # 10月以下の場合、前年の10月を生成
    if dt.month <= 10:
        prev_october = datetime(dt.year - 1, 10, 1)
    else:
        # 11月、12月の場合、今年の10月を生成
        prev_october = datetime(dt.year, 10, 1)
    return prev_october.strftime("%Y%m")

# データ件数に応じて行を追加
def _insert_rows_with_style(ws, base_row: int, data_count: int, isCopyValue: bool):
    summary_base = ws[base_row]
    ws.insert_rows(base_row + 1, data_count - 2)
    for r in range(data_count - 2):
        i = 1
        for cell in summary_base:
            ws.cell(row = r + base_row + 1, column = i)._style = copy(cell._style)  # セルのスタイルをコピー
            if i == 2 and isCopyValue:
                ws.cell(row = r + base_row + 1, column = i).value = cell.value  # セルの値をコピー
            i += 1

# AzureIDに紐付くIDを取得
def get_user_id(conn, azure_ad_id: str):
    # 対応するユーザーID。存在しない場合は None。
    with conn.cursor() as cur:
        cur.execute("select id from m_users where azure_ad_id = %s", (azure_ad_id,))
        row = cur.fetchone()
        return row[0] if row else None

# デバッグ用サーバー起動
if __name__ == "__main__":
    app.run(debug=True,host='0.0.0.0',port=5000)
